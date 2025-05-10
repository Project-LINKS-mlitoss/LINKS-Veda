import asyncio
import itertools
import math

import pandas as pd
from openpyxl.reader.excel import load_workbook


async def fetch_excel(file_path):
    # Read with pandas
    xls = pd.ExcelFile(file_path, engine="openpyxl")

    # Read with openpyxl
    wb = load_workbook(file_path, data_only=True)

    return xls, wb


def fill_merged_na(sheet, dataframe):
    cells = sheet.merged_cells
    max = 0
    for e in cells.ranges:
        min_row, min_col, max_row, max_col = e.min_row, e.min_col, e.max_row, e.max_col
        if max_row + 1 - min_row > max:
            max = max_row + 1 - min_row
        base_value = sheet.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                dataframe.iloc[row - 1, col - 1] = base_value
    return dataframe, max


def find_header_row(df):
    index_header_rows = []
    for i in range(len(df) - 1):
        num_not_na = df.iloc[i].notna().sum()
        if num_not_na >= df.shape[1] * 0.7:
            string_column_count = df.iloc[i].apply(lambda x: isinstance(x, str)).sum()
            total_column = df.shape[1] * 0.89
            if string_column_count >= total_column:
                index_header_rows.append(i)  # Determine the header row
    if index_header_rows and len(index_header_rows) >= len(df) * 0.8:
        index_header_rows = index_header_rows[:1]
    if not index_header_rows:
        index_header_rows.append(0)
    return index_header_rows


def keep_last_of_consecutive(index_header_rows, max_row=0):
    # Sort the indexes and convert them into a list.
    sorted_rows = sorted(index_header_rows)

    # The result contains the last indexes of each consecutive sequence.
    result = []
    count = 0
    for i in range(len(sorted_rows)):
        # If it is the last element or not consecutive with the previous element.
        if i == len(sorted_rows) - 1 or sorted_rows[i] + 1 != sorted_rows[i + 1]:
            count += 1
            result.append(sorted_rows[i])
            if count > max_row:
                max_row = count
            count = 0
        else:
            count += 1
    return result, max_row


def fast_split_list(lst, indexes, max_row):
    indexes = sorted(set(indexes))  # Sort and remove duplicate indexes.
    result = []
    header_offset = []
    for i, j in zip([0] + indexes, indexes + [len(lst)]):
        if j < max_row:
            continue
        if i < j:
            e = i - max_row if i - max_row >= 0 else 0
            result.append(lst[e:j])
            header_offset.append(i - e)
    return result, header_offset


def chunk_list_with_header(lines, offset=0, keys_num=50):
    max_len = len(max(lines[offset:], key=len))
    chunk_size = math.ceil(1000 / max_len)
    chunk_size = max(min(chunk_size, math.floor(keys_num * 6 / 50)), 1)
    if not lines:
        return []
    header = lines[0:offset + 1]  # Take the first row as the header.
    print("*" * 20 + "HEADER")
    print(header)
    header = [h.strip().replace("\n", " ") for h in header]
    lines = lines[offset + 1:]
    chunks = ["\n".join(header + lines[i: i + chunk_size]) for i in range(0, len(lines), chunk_size)]
    for chunk in chunks:
        print("*" * 50)
        print(chunk)
    return chunks


async def chunk_excel_file(file_url, keys_num):
    xls, wb = await fetch_excel(file_url)
    final_chunk = []
    for sheet_name in xls.sheet_names:
        sheet = wb[sheet_name]
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        df, max_row = fill_merged_na(sheet, df)
        index_header_rows = sorted(set(find_header_row(df)))
        index_header_rows, max_row = keep_last_of_consecutive(index_header_rows, max_row)
        lines = df.fillna("").apply(lambda row: "|".join(row.astype(str)), axis=1)
        tables_chunking, header_offset = fast_split_list(lines.tolist(), index_header_rows, max_row)

        for table, offset in zip(tables_chunking, header_offset):
            table_chunks = chunk_list_with_header(table, offset, keys_num)
            final_chunk.extend(table_chunks)
    return final_chunk

from nlp.extraction import aextract_json_with_schema
from botocore.exceptions import ClientError

async def wrap_process(data, pydantic_model, additional_prompt, file_url):
    print("LLM Structuring excel record")
    result = []
    countRs = 1
    while True:
        try:
            obj = await aextract_json_with_schema("anthropic.claude-3-haiku-20240307-v1:0", {}, additional_prompt,
                                                  data, 'default', pydantic_model)
            llm_data = obj.get('llm_response_array', [])
            result.extend(llm_data)
            break
        except ClientError as e:
            error_code = e.response.get('Error', {}).get('Code', '')
            if error_code in ['ThrottlingException', 'TooManyRequestsException']:
                print("[llm_doc_structure] [Throttling Error] Waiting 60 seconds before retrying...")
                await asyncio.sleep(60)
            elif error_code == 'ValidationError':
                print(f"[llm_doc_structure] [Validation Error] Skip chunk of file {file_url}")
                break
            else:
                print(
                    f'[llm_doc_structure] [Unexpected ClientError] {e} when structure chunk file {file_url} at times {countRs}th')
                countRs += 1
                if countRs > 5:
                    break
                await asyncio.sleep(60)
        except Exception as e:
            print(f"Error occurred when structure chunk of file {file_url} at times {countRs}th")
            countRs += 1
            if countRs > 5:
                break
            await asyncio.sleep(60)
    return result

async def structure_excel_data(data, pydantic_model, additional_prompt, file_url):
    result = []
    for i in range(0,len(data), 10):
        batch = data[i:i + 10]
        arr_data = await asyncio.gather(*[wrap_process(record, pydantic_model, additional_prompt, file_url) for record in batch])
        result.extend(itertools.chain.from_iterable(arr_data))
    return result

